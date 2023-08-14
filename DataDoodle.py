# Project name :- DataDoodle
# Developer name :- Saurav Pandey
# Version : 1.6v

import pyautogui as auto
import openpyxl
from openpyxl import load_workbook
import sys

def Taking_TextFile_Path():
    try:
        global Text_File_Path
        Text_File_Path = auto.prompt("Path Of Text File","Input Here...").strip(''' "',.''')
        if Text_File_Path:
            if Text_File_Path.__contains__('.txt') :
                Text_File_Path = Text_File_Path.replace("\\", "\\\\")
                TakingExcelSheet_Path()
            else :
                auto.alert("\tInvalid path\t ","Try Again..")
                Taking_TextFile_Path()
        else:
            auto.alert("\tPath cannot be empty\t ","Try Again..")
            Taking_TextFile_Path()
    except:
        sys.exit()

def TakingExcelSheet_Path():
    try:
        global Excel_Sheet_Path
        Excel_Sheet_Path = auto.prompt("Path Of Excel Sheet","Input Here...").strip(''' "',.''')
        if Excel_Sheet_Path:
            if Excel_Sheet_Path.__contains__('.xlsx') :
                Excel_Sheet_Path = Excel_Sheet_Path.replace("\\", "\\\\")
                DeleteAndCreateSheet()
            else :
                auto.alert("\tInvalid path\t ","Try Again..")
                TakingExcelSheet_Path()
        else:
            auto.alert("\tPath cannot be empty\t ","Try Again..")
            TakingExcelSheet_Path()
    except:
        sys.exit()

def DeleteAndCreateSheet():
    try :
        # Open the Excel file
        workbook = load_workbook(Excel_Sheet_Path)
        # Delete all existing sheets
        for sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
        # Create a new sheet with a specific name
        workbook.create_sheet(title='Detailed Data')
        workbook.create_sheet(title='RewardsNumber')
        # Save the modified Excel file
        workbook.save(Excel_Sheet_Path)
        DetailedData()
    except : 
        auto.alert("\tError in Excel file.\t","Try Again..")
        TakingExcelSheet_Path()

def DetailedData():
    try:
        Data_List = []
        global rewardsNumber, errorReason, processedRecords
        rewardsNumber, errorReason, processedRecords = [],[],[]

        with open(Text_File_Path, 'r',errors='ignore',encoding='utf8') as MemberData:
            for line in MemberData:
                try:
                    Required_Data = line[line.find("reportOutput"):]
                    for i in range (Required_Data.count("jobId")):
                        if '}' in Required_Data:
                            Brace_End = Required_Data.find('}')+1
                        elif "[truncated]" in Required_Data:
                            Brace_End =  Required_Data.find("[truncated]")

                        One_Brace = Required_Data[Required_Data.find("{"):Brace_End].strip(" ")

                        Actual_Job_Id = Job_ID(One_Brace)
                        Actual_Object = Entity_Type(One_Brace)
                        Actual_Operation = Operations(One_Brace)
                        Actual_Processed_Records = Processed_Records(One_Brace)
                        Actual_Failed_Records = Failed_Records(One_Brace)
                        Actual_Error_Reason = Error_Reason(One_Brace)

                        Data_List.append([Actual_Job_Id,Actual_Object,Actual_Operation,Actual_Processed_Records,Actual_Failed_Records,Actual_Error_Reason])

                        Required_Data = Required_Data[Brace_End:]
                except:
                    auto.alert("\tError in Data.\t","Try Again..")
                    pass
            try :
                excelFile = openpyxl.load_workbook(Excel_Sheet_Path)
                detailedData = excelFile["Detailed Data"]
                detailedData.append(["Job Id","Object","Operation","Processed Records","Failed Records","Error Reason"])
                for data in Data_List:
                    detailedData.append(data)

                reportOutput = excelFile["RewardsNumber"]
                reportOutput.append(["Error Reason","Rewards Number","Processed Records"])

                for i in range(len(rewardsNumber)) :
                    Excel_Col = reportOutput['A' + str(i+2)]
                    Excel_Col.value = errorReason[i]

                    Excel_Col2 = reportOutput['B' + str(i+2)]
                    Excel_Col2.value = rewardsNumber[i]

                Excel_Col3 = reportOutput['C2']
                Excel_Col3.value = sum(processedRecords)

                excelFile.save(Excel_Sheet_Path)
            except:
                auto.alert("\tError in Excel file.\t","Try Again..")
                TakingExcelSheet_Path()

            auto.alert("\tDone ✅ \t ","Successful..")
    except:
        auto.alert("\tAn unexpected error Occured\t","Failed..")
        Taking_TextFile_Path()

def Job_ID(One_Brace):
    if "jobId" in One_Brace:
        try:
            Actual_Job_id = One_Brace[One_Brace.find("jobId")+5:One_Brace.find("jobId")+30].strip(''',. : '".,''')
            return Actual_Job_id
        except:
            return "Null"
    else:
        return None

def Entity_Type(One_Brace):
    if "entityType" in One_Brace:
        indexNo = One_Brace.find("entityType")
        try:
            Object = One_Brace[indexNo+10:indexNo+33].strip(''',. : '".,''')
            if "ContactMemberMaster" in Object:
                Actual_Object = "ContactMemberMaster__c"
            elif "Contact" in Object:
                Actual_Object = "Contact"
            return Actual_Object
        except:
            return "Null"
    else:
        return "Null"

def Operations(One_Brace):
    if "operation" in One_Brace:
        indexNo = One_Brace.find("operation")
        try:
            Actual_Operation = One_Brace[indexNo+13:indexNo+24]
            Actual_Operation = Actual_Operation[:Actual_Operation.find('''"''')].strip(''',. : '".,''')
            return Actual_Operation
        except:
            return "Null"
    else:
        return "Null"

def Processed_Records(One_Brace):
    if "processedRecords" in One_Brace:
        indexNo = One_Brace.find("processedRecords")
        try:
            Actual_Processed_Records = One_Brace[indexNo+16:indexNo+28].strip('''!~`@#$%^&*(-_+=[{)}]|\;></?,. : "'abcdefghijklmnopqrstuvwxyx.,ABCDEFGHIJKLMNOPQRSTUVWXYZ''')

            if(Actual_Processed_Records.isdigit()) :
                processedRecords.append(int(Actual_Processed_Records))
            return int(Actual_Processed_Records)
        except:
            return "Null"
    else:
        return "Null"
    
def Failed_Records(One_Brace):
    if "failedRecords" in One_Brace:
        indexNo = One_Brace.find("failedRecords")
        try:
            Actual_Failed_Records = One_Brace[indexNo+16:indexNo+20].strip('''!~`@#$%^&*(-_+=[{)}]|\;></?,. : "'abcdefghijklmnopqrstuvwxyx.,ABCDEFGHIJKLMNOPQRSTUVWXYZ''')
            return int(Actual_Failed_Records)
        except:
            return "Null"
    else:
        return "Null"

def Error_Reason(One_Brace):
    if "errorsList" in One_Brace:
        One_Brace_Copy = One_Brace
        try:
            Actual_Error_Reason = One_Brace[One_Brace.find("errorsList")+10:One_Brace.find('''"}''')].strip(''',. : '".,''')
    
            if Actual_Error_Reason == "No errors":
                return Actual_Error_Reason
            else:
                if "RewardsNumber" in One_Brace:
                    Reward_Number_Count = One_Brace.count("RewardsNumber")
                    Real_Error_Reason = ""
                    for i in range(Reward_Number_Count):
                        indexNo = One_Brace.find("RewardsNumber")
                        Rewards_Number = One_Brace[indexNo+13:indexNo+28].strip(",. : .,")

                        rewardsNumber.append(Rewards_Number)

                        One_Brace = One_Brace[One_Brace.find(Rewards_Number):]
    
                        if "--" in One_Brace:
                            error_Reason = One_Brace[0:One_Brace.find("--")+2].strip(""" 1234567890,.;:}[]{\|!@#$%^&*(+=)""")
                            errorReason.append(error_Reason)
                            if i == Reward_Number_Count - 1 :
                                Real_Error_Reason = Real_Error_Reason + error_Reason
                            else :
                                Real_Error_Reason = Real_Error_Reason + error_Reason + "\n"
                        else :
                            error_Reason = One_Brace[0:].strip(""" 1234567890,.;:}[]{\|!@#$%^&*(+=)""")
                            errorReason.append(error_Reason)
                            if i == Reward_Number_Count - 1 :
                                Real_Error_Reason = Real_Error_Reason + error_Reason
                            else :
                                Real_Error_Reason = Real_Error_Reason + error_Reason + "\n"
                    return Real_Error_Reason
        except:
            return "Null"
            
    elif "errorIfAny" in One_Brace :
        try:
            One_Brace = One_Brace[One_Brace.find("errorIfAny"):]
            Actual_Error_Reason = One_Brace[10:One_Brace.find(''',"''')].strip(''',. : '".,''')
            return Actual_Error_Reason 
        except:
            return "Null"
    else:
        return "Null"

auto.alert("\tWelcome\t\t","Pandey-Ji Creation..")
Taking_TextFile_Path()